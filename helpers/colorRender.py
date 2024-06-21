from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import numpy as np
import csv
import os

from searchBase import solver as S

def csv_to_numpy_matrix(file):
    with open(f'{file}', 'r') as csvfile:
        reader = csv.reader(csvfile)
        data = list(reader)
    
    numpy_matrix = np.array(data, dtype=int)
    return numpy_matrix

def bestSolution(dungeon):
    solutions = S.solve(dungeon)
    if len(solutions) > 0:
        best = solutions[0]
        for sol in solutions:
            if len(sol) < len(best):
                best = sol
    return best
def createFolder(nombre_carpeta):
    if not os.path.exists(nombre_carpeta):
        os.makedirs(nombre_carpeta)

def getFileNumber(nombre_carpeta):
    if not os.path.exists(nombre_carpeta):
        return 1
    else:
        return len(os.listdir(nombre_carpeta)) + 1

def createFoldersCsv(map):
    now = datetime.now()
    dateFormat = now.strftime("%d-%m")

    ExperimentFolder = "Results/Experiment " + dateFormat
    solutionsFolder = os.path.join(ExperimentFolder,"SolutionsCsv")

    createFolder(ExperimentFolder)
    createFolder(solutionsFolder)

    nFile = getFileNumber(solutionsFolder)
    csvName = f"Solution_{nFile}.csv"
    csvFileName = os.path.join(solutionsFolder, csvName)

    np.savetxt(f'{csvFileName}', map, delimiter=',', fmt='%d')
    return csvFileName

def createFoldersExcel(wb):
    now = datetime.now()
    dateFormat = now.strftime("%d-%m")

    ExperimentFolder = "Results/Experiment " + dateFormat
    dungeonsFolder = os.path.join(ExperimentFolder, "Dungeons")

    createFolder(ExperimentFolder)
    createFolder(dungeonsFolder)

    nFile= getFileNumber(dungeonsFolder)
    dungeonName = f"Dungeon_{nFile}.xlsx"
    excelFileName = os.path.join(dungeonsFolder,dungeonName)

    wb.save(f"{excelFileName}")

def fixColumWidth(worksheet,middleColumn):
    i=1
    for column_cells in worksheet.columns:
        max_length = 0
        column = column_cells[0].column_letter  
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        if middleColumn <= i:
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
        i+=1
 
def renderMatrix(map, fitness, nSol, minMoves):
    wb = Workbook()
    sheet = wb.active
    for fila_idx, fila in enumerate(map, start=1):
        for tile_idx, tileValue in enumerate(fila, start=1):
            sheet.cell(row=fila_idx, column=tile_idx, value=tileValue)
            if tileValue == 0:
                color = "F0F0F0"
            elif tileValue == 1:
                color = "C0C0C0"
            elif tileValue == 2:
                color = "D2B48C"
            elif tileValue == 3:
                color = "FFB6C1"
            elif tileValue == 4:
                color = "98FB98"
            elif tileValue == 5:
                color = "ADD8E6"
            filler = PatternFill(start_color=color, end_color=color, fill_type="solid")
            sheet.cell(row=fila_idx, column=tile_idx).fill = filler


   
    lastRow = (map.shape[0] // 2) - 1
    middleColumn = map.shape[1] + 2

    sheet.cell(row=lastRow, column=middleColumn,value="Fitness")
    sheet.cell(row=lastRow , column=middleColumn +1,value="Number of Solutions")
    sheet.cell(row=lastRow , column=middleColumn +2,value="Minimum Moves")

    lastRow += 1
    sheet.cell(row=lastRow, column=middleColumn,value=fitness)
    sheet.cell(row=lastRow , column=middleColumn +1,value=nSol)
    sheet.cell(row=lastRow , column=middleColumn +2,value=minMoves)

    fixColumWidth(sheet,middleColumn)
   

    createFoldersExcel(wb)
    createFoldersCsv(map)
    renderMatrixQueueWMoves(map)

def bestSolutionWMoves(dungeon):
    best =dungeon.pop()
    while dungeon:
        dungeono=dungeon.pop()
        if len(dungeono) < len(best):
            best = dungeono
    return best

def renderMatrixQueueWMoves(dungeon):
    dungeonWMoves=S.solveGameWithMoves(dungeon)
    matrix_queue = bestSolutionWMoves(dungeonWMoves)
    wb = Workbook()
    sheet = wb.active

    startColumn = 1
    while matrix_queue:
        map = matrix_queue.pop()
        lastRow = map.dungeon.shape[1] + 2 
        middleColumn = (map.dungeon.shape[0] // 2) + 1
        for fila_idx, fila in enumerate(map.dungeon, start=1):
            for tile_idx, tileValue in enumerate(fila, start=1):
                sheet.cell(row=fila_idx, column=tile_idx + startColumn - 1, value=tileValue)
                if tileValue == 0:
                    color = "F0F0F0"
                elif tileValue == 1:
                    color = "C0C0C0"
                elif tileValue == 2:
                    color = "D2B48C"
                elif tileValue == 3:
                    color = "FFB6C1"
                elif tileValue == 4:
                    color = "98FB98"
                elif tileValue == 5:
                    color = "ADD8E6"
                filler = PatternFill(start_color=color, end_color=color, fill_type="solid")
                sheet.cell(row=fila_idx, column=tile_idx + startColumn - 1).fill = filler

        if startColumn == 1: 
            sheet.cell(row=lastRow, column=middleColumn,value=map.move)
        startColumn += len(map.dungeon[0]) + 1
        if startColumn != 1:
            sheet.cell(row=lastRow, column=startColumn - middleColumn,value=map.move)


    now = datetime.now()
    dateFormat = now.strftime("%d-%m")
    ExperimentFolder = "Results/Experiment " + dateFormat
    solutionsFolder = os.path.join(ExperimentFolder,"Solutions")

    createFolder(solutionsFolder)

    nFile= getFileNumber(solutionsFolder)
    dungeonName = f"Solution_{nFile}.xlsx"
    excelFileName = os.path.join(solutionsFolder,dungeonName)

    wb.save(f"{excelFileName}")