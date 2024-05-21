from openpyxl import Workbook
from openpyxl.styles import PatternFill
from datetime import datetime
import numpy as np
import csv
from modules import solver as S

def csv_to_numpy_matrix(file):
    with open(f'results/{file}.csv', 'r') as csvfile:
        reader = csv.reader(csvfile)
        data = list(reader)
    
    # Convertimos los datos leídos a un array de NumPy
    numpy_matrix = np.array(data, dtype=int)
    return numpy_matrix

def bestSolution(dungeon):
    solutions = S.solve(dungeon)
    if len(solutions) > 0:
        best = solutions[0]
        for sol in solutions:
            if len(sol) < len(best):
                best = sol
    return best

def renderMatrix(map):
    wb = Workbook()
    hoja = wb.active
    for fila_idx, fila in enumerate(map, start=1):
        for celda_idx, valor_celda in enumerate(fila, start=1):
            hoja.cell(row=fila_idx, column=celda_idx, value=valor_celda)
            if valor_celda == 0:
                color = "F0F0F0"
            elif valor_celda == 1:
                color = "C0C0C0"
            elif valor_celda == 2:
                color = "D2B48C"
            elif valor_celda == 3:
                color = "FFB6C1"
            elif valor_celda == 4:
                color = "98FB98"
            elif valor_celda == 5:
                color = "ADD8E6"
            relleno = PatternFill(start_color=color, end_color=color, fill_type="solid")
            hoja.cell(row=fila_idx, column=celda_idx).fill = relleno

    now = datetime.now()
    formato_fecha = now.strftime("%d-%m_%H-%M-%S")
    nArchivo = "Dungeon_" + formato_fecha
    wb.save(f"results/{nArchivo}.xlsx")
    np.savetxt(f'results/{nArchivo}.csv', map, delimiter=',', fmt='%d')
    renderMatrixQueue(nArchivo)

def renderMatrixQueue(file):
    dungeon = csv_to_numpy_matrix(file)
    matrix_queue = bestSolution(dungeon)
    wb = Workbook()
    hoja = wb.active

    columna_inicio = 1
    while matrix_queue:
        map = matrix_queue.pop()
        for fila_idx, fila in enumerate(map, start=1):
            for celda_idx, valor_celda in enumerate(fila, start=1):
                hoja.cell(row=fila_idx, column=celda_idx + columna_inicio - 1, value=valor_celda)
                if valor_celda == 0:
                    color = "F0F0F0"
                elif valor_celda == 1:
                    color = "C0C0C0"
                elif valor_celda == 2:
                    color = "D2B48C"
                elif valor_celda == 3:
                    color = "FFB6C1"
                elif valor_celda == 4:
                    color = "98FB98"
                elif valor_celda == 5:
                    color = "ADD8E6"
                relleno = PatternFill(start_color=color, end_color=color, fill_type="solid")
                hoja.cell(row=fila_idx, column=celda_idx + columna_inicio - 1).fill = relleno

        columna_inicio += len(map[0]) + 1

    now = datetime.now()
    formato_fecha = now.strftime("%d-%m_%H-%M-%S")
    nArchivo = "Solution" + formato_fecha + ".xlsx"
    wb.save(f"results/{nArchivo}")